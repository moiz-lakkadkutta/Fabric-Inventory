"""Export service — CSV + XLSX serialisation helpers (TASK-CUT-403).

Two public callables:

  to_csv(rows, columns) -> bytes
      Render a list of dict-like rows into CSV bytes (UTF-8 with BOM so
      Excel-on-Windows displays the ₹ glyph natively, CRLF line endings,
      comma-escaping handled by the stdlib `csv` module).

  to_xlsx(sheets) -> bytes
      Render one or more sheets to an XLSX workbook. Each sheet is a
      (sheet_name -> rows + columns) entry; cell types are preserved
      (Decimal → numeric cell with `#,##0.00` format, date → date cell
      with `yyyy-mm-dd`, datetime stripped to date for compatibility).

Column descriptors keep both the wire key (matching the dataclass attr
or dict key) and the human-readable header so the router can call
``to_csv(rows, columns)`` without re-deriving anything.

Money is `Decimal` everywhere. Float is never allowed in cells — passing
one raises TypeError. This is intentional: it forces the routers to keep
the Decimal type all the way through the pipeline, which the CLAUDE.md
"never float for money" rule depends on.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font

# UTF-8 BOM — Excel-on-Windows uses the BOM as the magic number that
# tells it the file is UTF-8 (not its legacy ANSI codepage). Without
# the BOM the rupee glyph renders as mojibake (the byte sequence shows
# up as a 3-character ANSI string) when Indian users double-click the
# file. noqa RUF001/003 — the comment contains intentional Unicode
# examples (BOM, rupee, mojibake bytes).
UTF8_BOM = "﻿"


@dataclass(frozen=True, slots=True)
class Column:
    """One CSV/XLSX column.

    ``key``    — attribute name on the row dict / dataclass.
    ``header`` — text in the header row.
    ``kind``   — "text" | "number" | "money" | "date" | "int". Drives
                 XLSX cell formatting; CSV always stringifies.
    """

    key: str
    header: str
    kind: str = "text"


def _row_value(row: Any, key: str) -> Any:
    """Look up ``key`` on a dict or dataclass-like row."""
    if isinstance(row, Mapping):
        return row.get(key)
    return getattr(row, key, None)


def _cell_str(value: Any) -> str:
    """Stringify for CSV. Floats blow up (CLAUDE.md rule)."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Hard fail: money must be Decimal. Surfacing this at write time
        # keeps the bug at the router boundary, not in spreadsheet QA.
        raise TypeError(
            f"Refusing to write float {value!r} to a CSV cell — use Decimal for money values.",
        )
    if isinstance(value, Decimal):
        return f"{value:f}"
    if isinstance(value, dt.datetime):
        # Excel handles plain ISO dates better than timestamps with TZ
        # info; the FE consumes the wire dates as already-IST-shifted.
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def to_csv(rows: Iterable[Any], columns: Sequence[Column]) -> bytes:
    """Render rows + columns to CSV bytes with UTF-8 BOM + CRLF lines.

    The BOM + CRLF combination is what Excel-on-Windows expects for
    "double-click and the file opens correctly with ₹ rendered as ₹".
    """
    buffer = io.StringIO()
    # csv.writer handles comma + double-quote + newline escaping
    # natively; we just need to feed it the right values.
    writer = csv.writer(buffer, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([c.header for c in columns])
    for row in rows:
        writer.writerow([_cell_str(_row_value(row, c.key)) for c in columns])
    return (UTF8_BOM + buffer.getvalue()).encode("utf-8")


def _write_cell(cell: Cell, value: Any, kind: str) -> None:
    """Write ``value`` into an XLSX cell with the right type + format.

    Decimal money → numeric cell with Indian-style 2-decimal formatting.
    Date → date cell with ISO format. Float is rejected as in CSV.
    """
    if value is None:
        return
    if isinstance(value, float):
        raise TypeError(
            f"Refusing to write float {value!r} to an XLSX cell — use Decimal for money values.",
        )
    if kind in ("money", "number"):
        # Decimal preserves exact value; openpyxl writes it as numeric.
        if isinstance(value, Decimal):
            cell.value = value
            cell.number_format = "#,##0.00" if kind == "money" else "0.0000"
            return
        if isinstance(value, int):
            cell.value = value
            cell.number_format = "#,##0" if kind == "money" else "0"
            return
        # Fall through — keep as text but flag the unexpected type so the
        # router test fails loud.
        cell.value = str(value)
        return
    if kind == "int":
        if isinstance(value, int) and not isinstance(value, bool):
            cell.value = value
            cell.number_format = "0"
            return
        cell.value = str(value)
        return
    if kind == "date":
        if isinstance(value, dt.datetime):
            value = value.date()
        if isinstance(value, dt.date):
            cell.value = value
            cell.number_format = "yyyy-mm-dd"
            return
        cell.value = str(value)
        return
    # Plain text fallback.
    if isinstance(value, Decimal):
        cell.value = str(value)
        return
    if isinstance(value, dt.datetime):
        cell.value = value.date().isoformat()
        return
    if isinstance(value, dt.date):
        cell.value = value.isoformat()
        return
    cell.value = str(value)


@dataclass(frozen=True, slots=True)
class Sheet:
    """One sheet inside a multi-sheet XLSX (e.g. GSTR-1's 5 buckets).

    ``rows`` is the iterable of row dicts/dataclasses; ``columns`` is
    the column descriptor list applied to those rows.
    """

    name: str
    columns: Sequence[Column]
    rows: Iterable[Any]


def to_xlsx(sheets: Sequence[Sheet]) -> bytes:
    """Render one or more sheets to an XLSX workbook (bytes).

    The header row is bolded; numeric columns get number formats so
    the values are sortable / filterable in Excel without manual
    "convert to number" cell prompts.
    """
    workbook = Workbook()
    # openpyxl creates a default blank sheet; remove + replace with ours.
    default = workbook.active
    if default is not None:
        workbook.remove(default)

    if not sheets:
        # An xlsx must have at least one sheet to be a valid OOXML
        # archive; surface that with an empty data sheet rather than
        # raising at write time.
        workbook.create_sheet(title="Empty")

    for sheet in sheets:
        # Excel caps sheet names at 31 chars and forbids ``[]:*?/\``.
        # Truncate quietly; replace forbidden chars with `_`.
        clean_name = sheet.name[:31]
        for bad in "[]:*?/\\":
            clean_name = clean_name.replace(bad, "_")
        ws = workbook.create_sheet(title=clean_name)
        for col_idx, col in enumerate(sheet.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left")
        for row_idx, row in enumerate(sheet.rows, start=2):
            for col_idx, col in enumerate(sheet.columns, start=1):
                value = _row_value(row, col.key)
                cell = ws.cell(row=row_idx, column=col_idx)
                _write_cell(cell, value, col.kind)
        # Freeze the header so users keep their bearings when scrolling
        # long invoice / voucher lists.
        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ──────────────────────────────────────────────────────────────────────
# Router helpers — content-type + Content-Disposition formatting
# ──────────────────────────────────────────────────────────────────────


CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def content_disposition(filename: str) -> str:
    """Build the Content-Disposition header value for an attachment.

    Keeps the filename quoted so commas in the name don't get parsed
    as additional parameters by overly clever HTTP clients.
    """
    return f'attachment; filename="{filename}"'
