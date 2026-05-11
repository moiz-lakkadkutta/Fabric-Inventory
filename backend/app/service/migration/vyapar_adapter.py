"""VyaparExcelAdapter — TASK-CUT-402 Wave-5 implementation.

Implements the ``MigrationAdapter`` Protocol against Vyapar's built-in
Excel export. Vyapar publishes two relevant exports for the v1 cutover:

- **Parties** sheet (Utilities → Export → Excel → Parties): one row per
  party with name / phone / email / GSTIN / state / opening balance +
  type (To Receive / To Pay).
- **Opening balances** sheet (same export, when balances exist): same
  party rows; we derive the OB rows directly from the parties sheet
  rather than expecting a separate sheet. Vyapar exports balances on
  the party rows, not on a dedicated sheet. (Future versions may
  split; the adapter ignores sheets it doesn't recognise.)

Per the Wave-1 spike (``docs/spikes/vyapar-source-format.md``) we
deliberately keep this thin and column-driven: known headers map to
intermediate fields, anything else is dropped on the floor. The
column-mapping lives in ``_COLUMN_MAP`` and can be extended without
touching the adapter logic — adding a Hindi-locale variant of "Name"
is one new tuple.

The adapter is purely a parser:

- Does NOT write to Postgres.
- Does NOT resolve foreign keys.
- Does NOT compute TB reconciliation (the commit step does, since the
  TB sides depend on the seeded COA which is per-org).

The wrapping ``migration_service`` handles all of that. Validation
errors (unparseable rows, GSTIN format) become ``ReconciliationRow``
entries that downstream tooling renders verbatim — the FE preview
pane is the user-facing surface.

Money parsing: Vyapar serializes amounts as either bare numbers
(``"1234.50"``) or with the rupee sign + Indian commas
(``"₹1,23,450.00"``). The ``_parse_decimal`` helper strips both and
parses through ``Decimal`` — never via float, per CLAUDE.md.
"""

from __future__ import annotations

import re
from collections.abc import Iterable, Iterator
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl

from .intermediate import (
    IntermediateOpeningBalance,
    IntermediateParty,
    MigrationValidationReport,
    PartyKind,
    ReconciliationRow,
)

# Source-format identifier — persisted in `user_migration.source_format`.
SOURCE_FORMAT = "vyapar_excel"

# Sheet name candidates. Vyapar's UI varies: "Parties", "Party List",
# localised "पार्टी". We accept any of the canonical English forms; a
# future i18n extension drops new entries here.
_PARTY_SHEET_NAMES = {"parties", "party", "party list", "party master", "customers", "suppliers"}

# Canonical column names → IntermediateParty field name. Lower-cased on
# both sides for case-insensitive matching. Multiple Vyapar localisations
# map onto the same canonical field.
_COLUMN_MAP: dict[str, str] = {
    # Required identity
    "name": "name",
    "party name": "name",
    "party": "name",
    "customer name": "name",
    "supplier name": "name",
    # Code is optional in Vyapar — derived from name if absent
    "code": "code",
    "party code": "code",
    "alias": "code",
    # Contact
    "contact": "contact_person",
    "contact person": "contact_person",
    "email": "email",
    "email id": "email",
    "phone": "phone",
    "phone number": "phone",
    "mobile": "phone",
    "mobile number": "phone",
    # Tax / state
    "gstin": "gstin",
    "gst number": "gstin",
    "gst no": "gstin",
    "pan": "pan",
    "pan number": "pan",
    "state": "state_code",
    "state code": "state_code",
    # Address
    "address": "address",
    "billing address": "address",
    # Opening balance — the magnitude. Sign / DR/CR comes from the
    # accompanying "type" column.
    "opening balance": "_opening_balance",
    "balance": "_opening_balance",
    "amount": "_opening_balance",
    # Balance type — "To Receive" (we owe money TO them == we are creditor;
    # they owe us == debtor) is Vyapar's wording. "To Pay" means we owe
    # them (they are a creditor of ours). Mapped during processing.
    "balance type": "_balance_type",
    "type": "_party_type",
    "party type": "_party_type",
}


# Vyapar party-type column values → kinds tuple. Defaults to CUSTOMER
# (most common starter party in a Vyapar firm) when the column is
# absent or empty.
_PARTY_TYPE_MAP: dict[str, tuple[PartyKind, ...]] = {
    "customer": ("CUSTOMER",),
    "customers": ("CUSTOMER",),
    "supplier": ("SUPPLIER",),
    "suppliers": ("SUPPLIER",),
    "vendor": ("SUPPLIER",),
    "karigar": ("KARIGAR",),
    "job worker": ("KARIGAR",),
    "transporter": ("TRANSPORTER",),
}


# Indian rupee sign + commas + whitespace, in any combination. Regex is
# intentionally permissive — we only need the digits and the optional
# leading minus + decimal point.
_RUPEE_NOISE_RE = re.compile(r"[₹,\s]")
_GSTIN_RE = re.compile(r"^\d{2}[A-Z]{5}\d{4}[A-Z][1-9A-Z]Z[A-Z\d]$")


def _parse_decimal(value: Any) -> Decimal:
    """Parse a money value to Decimal, tolerating Vyapar's quirks.

    Accepts ``None`` (returns ``Decimal('0')``), numeric (int/float —
    converted via str to dodge binary-float precision loss), or string
    with optional rupee sign / Indian comma grouping. Raises on
    unparseable input — callers wrap that into a reconciliation error.
    """
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, bool):
        # bools are ints in Python; reject explicitly so a stray TRUE/FALSE
        # cell doesn't quietly become 1/0.
        raise ValueError(f"Unexpected boolean opening balance: {value!r}")
    if isinstance(value, int | float):
        # Decimal(float) carries the float's binary noise. Always go via str.
        return Decimal(str(value))
    if not isinstance(value, str):
        raise ValueError(f"Unsupported opening-balance type: {type(value).__name__}")
    cleaned = _RUPEE_NOISE_RE.sub("", value).strip()
    if cleaned in ("", "-"):
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"Could not parse {value!r} as Decimal") from exc


def _normalize_str(value: Any) -> str | None:
    """Trim a cell value to a non-empty string, else None."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _derive_code(name: str, fallback_index: int) -> str:
    """Synthesize a party code from the name when the source has none.

    Strategy: take alphanumerics from the first three whitespace-separated
    tokens, upper-case, hyphenate, truncate to 50 chars. Falls back to
    ``PARTY-<index>`` if the name has no alphanumerics (e.g. punctuation
    only — exceedingly rare).
    """
    tokens = re.findall(r"[A-Za-z0-9]+", name)
    if not tokens:
        return f"PARTY-{fallback_index}"
    code = "-".join(t.upper() for t in tokens[:3])
    return code[:50] or f"PARTY-{fallback_index}"


def _resolve_party_sheet(workbook: Any) -> Any | None:
    """Return the sheet that holds the parties list, or None if absent."""
    for sheet in workbook.worksheets:
        if (sheet.title or "").strip().lower() in _PARTY_SHEET_NAMES:
            return sheet
    # Single-sheet exports happen too — fall back to the first sheet.
    if len(workbook.worksheets) == 1:
        return workbook.worksheets[0]
    return None


def _read_header(sheet: Any) -> dict[int, str]:
    """Return ``{column_index: canonical_field_name}`` for the header row.

    Unknown columns are skipped silently — the adapter doesn't choke on
    Vyapar adding new columns we don't care about. Multiple sheet
    headers mapping to the same field is fine; the last wins (rare).
    """
    header_map: dict[int, str] = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        return header_map
    for idx, cell in enumerate(header_row):
        text = _normalize_str(cell)
        if not text:
            continue
        canonical = _COLUMN_MAP.get(text.lower())
        if canonical is not None:
            header_map[idx] = canonical
    return header_map


def _load_workbook(source: Any) -> Any:
    """Open an Excel workbook from a path, bytes, or BytesIO."""
    if isinstance(source, str | Path):
        return openpyxl.load_workbook(str(source), data_only=True, read_only=True)
    if isinstance(source, bytes | bytearray | memoryview):
        return openpyxl.load_workbook(BytesIO(bytes(source)), data_only=True, read_only=True)
    if isinstance(source, BytesIO):
        source.seek(0)
        return openpyxl.load_workbook(source, data_only=True, read_only=True)
    # Last-ditch: openpyxl accepts file-like objects too.
    return openpyxl.load_workbook(source, data_only=True, read_only=True)


def _iter_party_rows(source: Any) -> Iterator[tuple[int, dict[str, Any]]]:
    """Yield ``(source_row_index, raw_dict)`` for every data row.

    ``source_row_index`` is 1-based (matches what a user sees in Excel
    — header is row 1, data starts at 2). Empty rows are skipped.
    """
    workbook = _load_workbook(source)
    sheet = _resolve_party_sheet(workbook)
    if sheet is None:
        return
    header = _read_header(sheet)
    if not header:
        return
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue  # blank row
        raw: dict[str, Any] = {}
        for col_idx, value in enumerate(row):
            field = header.get(col_idx)
            if field is None:
                continue
            raw[field] = value
        if not raw:
            continue
        # Skip rows with no name (Vyapar's export sometimes pads a
        # blank "Total" row at the bottom).
        if not _normalize_str(raw.get("name")):
            continue
        yield row_idx, raw


def _resolve_kinds(raw_type: Any, raw_balance_type: Any) -> tuple[PartyKind, ...]:
    """Map Vyapar's party-type / balance-type columns to canonical kinds.

    Precedence:
        - Explicit "type" column wins if present (CUSTOMER / SUPPLIER / …).
        - Otherwise infer from "balance type": "To Receive" → CUSTOMER,
          "To Pay" → SUPPLIER.
        - Default: CUSTOMER (Vyapar's default party kind).
    """
    type_text = _normalize_str(raw_type)
    if type_text:
        mapped = _PARTY_TYPE_MAP.get(type_text.lower())
        if mapped is not None:
            return mapped
    balance_type = _normalize_str(raw_balance_type)
    if balance_type:
        bt = balance_type.lower()
        if "receive" in bt:
            return ("CUSTOMER",)
        if "pay" in bt:
            return ("SUPPLIER",)
    return ("CUSTOMER",)


def _resolve_ob_side(kinds: tuple[PartyKind, ...], raw_balance_type: Any) -> str:
    """Pick DR or CR for a party-scoped opening balance.

    Vyapar's "To Receive" means the party owes us → debit the receivable
    (sundry debtors DR). "To Pay" means we owe them → credit the
    payable (sundry creditors CR). When the balance-type column is
    absent we fall back to party kind: CUSTOMER → DR (we receive),
    SUPPLIER → CR (we pay), else DR by default.
    """
    bt = _normalize_str(raw_balance_type)
    if bt:
        low = bt.lower()
        if "receive" in low:
            return "DR"
        if "pay" in low:
            return "CR"
    if "SUPPLIER" in kinds:
        return "CR"
    return "DR"


def _resolve_ob_ledger_kind(kinds: tuple[PartyKind, ...]) -> str:
    """Sundry Debtors for customers; Sundry Creditors for suppliers."""
    if "SUPPLIER" in kinds or "KARIGAR" in kinds:
        return "SUNDRY_CREDITORS"
    return "SUNDRY_DEBTORS"


class VyaparExcelAdapter:
    """Vyapar's Excel-export reader. Implements ``MigrationAdapter``.

    Adapters are stateless: each method re-opens the workbook. This
    keeps the surface trivially testable (pass a fresh fixture for each
    case) and matches the Protocol's "the source is whatever the
    adapter wants" contract.

    ``source`` accepts a path (``str`` / ``Path``), raw ``bytes``, or a
    ``BytesIO`` — the wrapping ``POST /admin/migrations`` endpoint
    hands in the upload's bytes.
    """

    source_format = SOURCE_FORMAT

    def extract_parties(self, source: Any) -> Iterable[IntermediateParty]:
        """Yield ``IntermediateParty`` rows for every parseable party.

        Errors at this layer are reserved for the ``validate`` pass —
        ``extract_parties`` skips rows it can't parse without raising,
        so the commit step has a stable lower-bound set of parties to
        work with even if validation flagged some warnings.
        """
        for source_idx, raw in _iter_party_rows(source):
            name = _normalize_str(raw.get("name"))
            if not name:
                continue
            kinds = _resolve_kinds(raw.get("_party_type"), raw.get("_balance_type"))
            code = _normalize_str(raw.get("code")) or _derive_code(name, source_idx)
            gstin = _normalize_str(raw.get("gstin"))
            # Pre-validate GSTIN format; otherwise the commit-step's
            # masters_service will reject the row at insert time.
            if gstin and not _GSTIN_RE.fullmatch(gstin):
                gstin = None
            state_code = _normalize_str(raw.get("state_code"))
            if state_code:
                # Vyapar sometimes writes "Maharashtra" or "27". Take the
                # leading 2 alnum chars if they look like a code; else
                # drop. Real state-code resolution is out of v1 scope.
                if state_code.isdigit() and len(state_code) <= 2:
                    state_code = state_code.zfill(2)
                elif len(state_code) == 2 and state_code.isalpha():
                    state_code = state_code.upper()
                else:
                    state_code = None
            pan = _normalize_str(raw.get("pan"))
            yield IntermediateParty(
                source_id=str(source_idx),
                name=name,
                code=code,
                kinds=kinds,
                gstin=gstin,
                pan=pan,
                state_code=state_code,
                contact_person=_normalize_str(raw.get("contact_person")),
                email=_normalize_str(raw.get("email")),
                phone=_normalize_str(raw.get("phone")),
                address=_normalize_str(raw.get("address")),
            )

    def extract_opening_balances(self, source: Any) -> Iterable[IntermediateOpeningBalance]:
        """Yield ``IntermediateOpeningBalance`` for every party with a non-zero balance.

        Each balance points back at the matching party via
        ``party_source_id`` (same value used as the party's
        ``source_id``). The commit step resolves the FK after creating
        the parties.

        Firm-level openings (cash / capital / bank) are NOT in scope for
        v1 — Vyapar's "Parties" export covers only party-scoped
        balances. A future v2 task can ingest the "Cash / Bank" sheet
        and emit firm-level rows.
        """
        for source_idx, raw in _iter_party_rows(source):
            try:
                amount = _parse_decimal(raw.get("_opening_balance"))
            except ValueError:
                continue  # validate() flags it; extract just skips
            if amount == 0:
                continue
            kinds = _resolve_kinds(raw.get("_party_type"), raw.get("_balance_type"))
            side = _resolve_ob_side(kinds, raw.get("_balance_type"))
            ledger_kind = _resolve_ob_ledger_kind(kinds)
            yield IntermediateOpeningBalance(
                source_id=f"OB-{source_idx}",
                party_source_id=str(source_idx),
                ledger_kind=ledger_kind,  # type: ignore[arg-type]  # Literal narrowed by _resolve_ob_ledger_kind
                amount=abs(amount),
                side=side,  # type: ignore[arg-type]  # Literal narrowed by _resolve_ob_side
                narration=f"Opening balance imported from Vyapar (row {source_idx})",
            )

    def validate(self, source: Any) -> MigrationValidationReport:
        """Non-destructive pass — collects parse errors / warnings.

        Caller (commit step) populates ``tb_reconciles`` + ``tb_diff``
        because TB reconciliation needs the seeded COA, which the
        adapter doesn't have.
        """
        rows: list[ReconciliationRow] = []
        total_parties = 0
        total_opening_balances = 0
        errors = 0
        warnings = 0

        seen_source_ids: set[str] = set()

        for source_idx, raw in _iter_party_rows(source):
            total_parties += 1
            source_ref = f"row:{source_idx}"
            name = _normalize_str(raw.get("name"))
            if not name:
                rows.append(
                    ReconciliationRow(
                        severity="error",
                        code="PARTY_NAME_MISSING",
                        message="Party row missing name; will be skipped on commit.",
                        source_ref=source_ref,
                    )
                )
                errors += 1
                continue

            source_id = str(source_idx)
            if source_id in seen_source_ids:
                # Defensive — _iter_party_rows yields a unique row index
                # per row by construction; this would only trip if the
                # source had duplicate Excel row numbers (impossible).
                rows.append(
                    ReconciliationRow(
                        severity="error",
                        code="DUPLICATE_SOURCE_ID",
                        message=f"Duplicate source row id {source_id!r}.",
                        source_ref=source_ref,
                    )
                )
                errors += 1
                continue
            seen_source_ids.add(source_id)

            gstin = _normalize_str(raw.get("gstin"))
            if gstin and not _GSTIN_RE.fullmatch(gstin):
                rows.append(
                    ReconciliationRow(
                        severity="warn",
                        code="GSTIN_FORMAT_INVALID",
                        message=(
                            f"GSTIN {gstin!r} on row {source_idx} doesn't match the standard "
                            "15-character format; party will be imported without GSTIN."
                        ),
                        source_ref=source_ref,
                    )
                )
                warnings += 1

            # Opening balance parse check + count
            ob_raw = raw.get("_opening_balance")
            if ob_raw not in (None, ""):
                try:
                    amount = _parse_decimal(ob_raw)
                except ValueError:
                    rows.append(
                        ReconciliationRow(
                            severity="error",
                            code="OPENING_BALANCE_UNPARSEABLE",
                            message=(
                                f"Could not parse opening balance {ob_raw!r} on row {source_idx}."
                            ),
                            source_ref=source_ref,
                        )
                    )
                    errors += 1
                else:
                    if amount != 0:
                        total_opening_balances += 1

        if total_parties == 0:
            rows.append(
                ReconciliationRow(
                    severity="error",
                    code="NO_PARTIES_FOUND",
                    message=(
                        "No party rows detected. Check the workbook has a sheet "
                        "named one of: Parties, Party List, Customers, Suppliers."
                    ),
                    source_ref=None,
                )
            )
            errors += 1
        else:
            rows.insert(
                0,
                ReconciliationRow(
                    severity="info",
                    code="EXTRACTED",
                    message=(
                        f"Extracted {total_parties} parties and "
                        f"{total_opening_balances} opening balances from Vyapar export."
                    ),
                    source_ref=None,
                ),
            )

        return MigrationValidationReport(
            total_parties=total_parties,
            total_opening_balances=total_opening_balances,
            errors=errors,
            warnings=warnings,
            rows=tuple(rows),
            tb_reconciles=None,
            tb_diff=None,
        )


__all__ = ["SOURCE_FORMAT", "VyaparExcelAdapter"]
