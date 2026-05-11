"""End-to-end banking export tests — TASK-CUT-501b.

Covers ``?format=csv|xlsx`` on the two banking list endpoints that
TASK-CUT-403 left out (called out in `docs/retros/task-CUT-403.md`):

* GET /bank-accounts  (CSV + XLSX)
* GET /cheques        (CSV + XLSX)

Each test asserts the Content-Type, the attachment header, and one
content invariant. RLS-isolation is inherited from the JSON list
endpoint (same RLS-scoped query path; verified at the service level
in `tests/test_banking_routers.py`).
"""

from __future__ import annotations

import io
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.engine import Engine

from app.service.export_service import UTF8_BOM


def _signup_owner(http_client: TestClient) -> dict[str, str]:
    """Sign up + switch firm so the JWT carries firm_id."""
    resp = http_client.post(
        "/auth/signup",
        json={
            "email": f"u-{uuid.uuid4().hex[:10]}@example.com",
            "password": "strong-password-1",
            "org_name": f"Org-{uuid.uuid4().hex[:8]}",
            "firm_name": "Primary",
            "state_code": "MH",
        },
    )
    assert resp.status_code == 201, resp.text
    body: dict[str, str] = resp.json()
    switch = http_client.post(
        "/auth/switch-firm",
        headers={"Authorization": f"Bearer {body['access_token']}"},
        json={"firm_id": body["firm_id"]},
    )
    assert switch.status_code == 200, switch.text
    body["access_token"] = switch.json()["access_token"]
    return body


def _seed_ledger(sync_engine: Engine, *, org_id: str, firm_id: str) -> str:
    with sync_engine.connect() as conn, conn.begin():
        conn.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        coa_group_id = conn.execute(
            text(
                "INSERT INTO coa_group (org_id, code, name, group_type) "
                "VALUES (:org_id, 'ASSET', 'Assets', 'ASSET') "
                "ON CONFLICT DO NOTHING "
                "RETURNING coa_group_id"
            ),
            {"org_id": org_id},
        ).scalar_one_or_none()
        if coa_group_id is None:
            coa_group_id = conn.execute(
                text(
                    "SELECT coa_group_id FROM coa_group WHERE org_id = :org_id AND code = 'ASSET'"
                ),
                {"org_id": org_id},
            ).scalar_one()
        ledger_id = conn.execute(
            text(
                "INSERT INTO ledger (org_id, firm_id, code, name, coa_group_id) "
                "VALUES (:org_id, :firm_id, :code, 'Main Bank', :coa_group_id) "
                "RETURNING ledger_id"
            ),
            {
                "org_id": org_id,
                "firm_id": firm_id,
                "code": f"BANK-{uuid.uuid4().hex[:6].upper()}",
                "coa_group_id": coa_group_id,
            },
        ).scalar_one()
    return str(ledger_id)


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _create_bank_account(
    http_client: TestClient,
    me: dict[str, str],
    ledger_id: str,
    *,
    bank_name: str = "HDFC Bank",
    account_number: str = "00123456789012",
) -> str:
    resp = http_client.post(
        "/bank-accounts",
        headers=_auth(me["access_token"]),
        json={
            "firm_id": me["firm_id"],
            "ledger_id": ledger_id,
            "bank_name": bank_name,
            "account_number": account_number,
            "ifsc_code": "HDFC0001234",
            "account_type": "CURRENT",
            "balance": "100000.00",
        },
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["bank_account_id"]


CSV_MEDIA = "text/csv"
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _assert_attachment(resp: object, media: str, suffix: str) -> None:
    assert resp.status_code == 200, resp.text  # type: ignore[attr-defined]
    assert resp.headers["content-type"].startswith(media), resp.headers["content-type"]  # type: ignore[attr-defined]
    disp = resp.headers["content-disposition"]  # type: ignore[attr-defined]
    assert disp.startswith('attachment; filename="'), disp
    assert disp.rstrip('"').endswith(f".{suffix}"), disp


# ──────────────────────────────────────────────────────────────────────
# Bank accounts
# ──────────────────────────────────────────────────────────────────────


def test_bank_accounts_csv_export_returns_attachment_with_bom(
    http_client: TestClient, sync_engine: Engine
) -> None:
    me = _signup_owner(http_client)
    ledger_id = _seed_ledger(sync_engine, org_id=me["org_id"], firm_id=me["firm_id"])
    _create_bank_account(http_client, me, ledger_id, bank_name="HDFC Bank Export")

    resp = http_client.get("/bank-accounts?format=csv", headers=_auth(me["access_token"]))
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert text_body.startswith(UTF8_BOM), "Excel-on-Windows needs UTF-8 BOM"
    # Header row + at least one data row.
    assert "Bank,Account #,IFSC" in text_body
    assert "HDFC Bank Export" in text_body
    # PII (account number) is decrypted plaintext in the export.
    assert "00123456789012" in text_body


def test_bank_accounts_xlsx_export_returns_workbook(
    http_client: TestClient, sync_engine: Engine
) -> None:
    me = _signup_owner(http_client)
    ledger_id = _seed_ledger(sync_engine, org_id=me["org_id"], firm_id=me["firm_id"])
    _create_bank_account(http_client, me, ledger_id, bank_name="ICICI XLSX Test")

    resp = http_client.get("/bank-accounts?format=xlsx", headers=_auth(me["access_token"]))
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Bank accounts" in wb.sheetnames
    ws = wb["Bank accounts"]
    headers = [c.value for c in ws[1]]
    assert "Bank" in headers
    assert "Balance" in headers
    bank_col = headers.index("Bank") + 1
    bank_cells = [ws.cell(row=r, column=bank_col).value for r in range(2, ws.max_row + 1)]
    assert "ICICI XLSX Test" in bank_cells
    # Money cells get the Indian-style 2-decimal number format.
    balance_col = headers.index("Balance") + 1
    balance_cell = ws.cell(row=2, column=balance_col)
    assert "0.00" in (balance_cell.number_format or "")


# ──────────────────────────────────────────────────────────────────────
# Cheques
# ──────────────────────────────────────────────────────────────────────


def test_cheques_csv_export_returns_attachment(
    http_client: TestClient, sync_engine: Engine
) -> None:
    me = _signup_owner(http_client)
    ledger_id = _seed_ledger(sync_engine, org_id=me["org_id"], firm_id=me["firm_id"])
    bank_account_id = _create_bank_account(http_client, me, ledger_id)

    rcpt = http_client.post(
        f"/cheques?firm_id={me['firm_id']}",
        headers=_auth(me["access_token"]),
        json={
            "bank_account_id": bank_account_id,
            "cheque_number": "EXP001",
            "cheque_date": "2026-04-27",
            "payee_name": "Karigar Co",
            "amount": "5000.00",
        },
    )
    assert rcpt.status_code == 201, rcpt.text

    resp = http_client.get(
        f"/cheques?bank_account_id={bank_account_id}&format=csv",
        headers=_auth(me["access_token"]),
    )
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert text_body.startswith(UTF8_BOM)
    assert "Cheque #,Date,Payee" in text_body
    assert "EXP001" in text_body
    assert "Karigar Co" in text_body
    assert "5000.00" in text_body


def test_cheques_xlsx_export_returns_workbook(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    ledger_id = _seed_ledger(sync_engine, org_id=me["org_id"], firm_id=me["firm_id"])
    bank_account_id = _create_bank_account(http_client, me, ledger_id)

    http_client.post(
        f"/cheques?firm_id={me['firm_id']}",
        headers=_auth(me["access_token"]),
        json={
            "bank_account_id": bank_account_id,
            "cheque_number": "EXP002",
            "cheque_date": "2026-04-28",
            "payee_name": "Supplier XLSX",
            "amount": "12500.00",
        },
    )

    resp = http_client.get(
        f"/cheques?bank_account_id={bank_account_id}&format=xlsx",
        headers=_auth(me["access_token"]),
    )
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Cheques" in wb.sheetnames
    ws = wb["Cheques"]
    headers = [c.value for c in ws[1]]
    assert "Cheque #" in headers
    assert "Amount" in headers
    amount_col = headers.index("Amount") + 1
    amount_cell = ws.cell(row=2, column=amount_col)
    assert amount_cell.value is not None
    assert "0.00" in (amount_cell.number_format or "")


# ──────────────────────────────────────────────────────────────────────
# Auth-required
# ──────────────────────────────────────────────────────────────────────


def test_bank_accounts_export_requires_auth(http_client: TestClient) -> None:
    resp = http_client.get("/bank-accounts?format=csv")
    assert resp.status_code == 401


def test_cheques_export_requires_auth(http_client: TestClient) -> None:
    resp = http_client.get(f"/cheques?bank_account_id={uuid.uuid4()}&format=csv")
    assert resp.status_code == 401
