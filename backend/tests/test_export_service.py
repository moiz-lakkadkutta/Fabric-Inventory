"""Unit tests for the CSV/XLSX export helpers (TASK-CUT-403).

Service-only tests — no DB, no HTTP. Validates the contract the
routers depend on:

* CSV uses UTF-8 BOM + CRLF line endings + comma-escaped values so
  Excel-on-Windows displays ₹ / commas without prompting.
* XLSX preserves cell types (Decimal → numeric, date → date) and
  formats — money cells are `#,##0.00`, date cells `yyyy-mm-dd`.
* Float-as-money is rejected at write time (CLAUDE.md hard rule).
* Multi-sheet workbooks survive `openpyxl.load_workbook` round-trip.
"""

from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.service.export_service import (
    CSV_MEDIA_TYPE,
    UTF8_BOM,
    XLSX_MEDIA_TYPE,
    Column,
    Sheet,
    content_disposition,
    to_csv,
    to_xlsx,
)


def test_csv_has_utf8_bom_and_crlf_line_endings() -> None:
    columns = [Column("name", "Name"), Column("amount", "Amount", "money")]
    rows = [{"name": "Anjali", "amount": Decimal("1050.00")}]
    out = to_csv(rows, columns)
    text = out.decode("utf-8")
    assert text.startswith(UTF8_BOM), "Excel-on-Windows needs the BOM"
    assert "\r\n" in text, "Excel needs CRLF, not bare \\n"
    # Header + one row + trailing CRLF on the row.
    assert text.split("\r\n")[0] == UTF8_BOM + "Name,Amount"
    assert text.split("\r\n")[1] == "Anjali,1050.00"


def test_csv_escapes_embedded_commas_and_quotes() -> None:
    columns = [Column("name", "Name"), Column("notes", "Notes")]
    rows = [{"name": "Acme, Inc.", "notes": 'He said "hi"'}]
    out = to_csv(rows, columns).decode("utf-8")
    # csv.writer quotes the comma-bearing field; embedded quotes are
    # doubled per RFC 4180.
    assert '"Acme, Inc."' in out
    assert '"He said ""hi"""' in out


def test_csv_rejects_float_money() -> None:
    """Float for money is a CLAUDE.md hard rule — should raise."""
    columns = [Column("amount", "Amount", "money")]
    rows = [{"amount": 1050.50}]
    with pytest.raises(TypeError, match="float"):
        to_csv(rows, columns)


def test_csv_renders_rupee_glyph_in_utf8() -> None:
    columns = [Column("label", "Label")]
    rows = [{"label": "Total ₹1,050"}]
    out = to_csv(rows, columns)
    # ₹ is U+20B9 — only round-trips as 3 bytes when written as UTF-8.
    assert "₹".encode() in out


def test_xlsx_round_trips_decimal_as_number_cell() -> None:
    """openpyxl writes Decimal as a numeric cell; the round-trip read
    back gives us a number, not a string. This is the load-bearing
    promise for users who pivot/sum/sort in Excel.
    """
    sheet = Sheet(
        name="Invoices",
        columns=[Column("number", "Number"), Column("amount", "Amount", "money")],
        rows=[{"number": "RT/2526/0001", "amount": Decimal("1050.00")}],
    )
    raw = to_xlsx([sheet])
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Invoices"]
    assert ws["A1"].value == "Number"
    assert ws["B1"].value == "Amount"
    # The data cell should be numeric (Decimal/float), not a string.
    assert ws["B2"].value == Decimal("1050.00") or ws["B2"].value == 1050.00
    assert "0.00" in (ws["B2"].number_format or "")


def test_xlsx_formats_dates_as_iso() -> None:
    sheet = Sheet(
        name="Receipts",
        columns=[Column("date", "Date", "date")],
        rows=[{"date": dt.date(2026, 4, 30)}],
    )
    raw = to_xlsx([sheet])
    wb = load_workbook(io.BytesIO(raw))
    cell = wb["Receipts"]["A2"]
    # openpyxl reads date cells back as Python datetime in modern versions.
    assert cell.value in (dt.date(2026, 4, 30), dt.datetime(2026, 4, 30))
    assert (cell.number_format or "").lower().startswith("yyyy")


def test_xlsx_supports_multiple_sheets_in_order() -> None:
    """GSTR-1's 5 buckets must each be its own sheet (B2B / B2CL / B2CS
    / Export / HSN)."""
    sheets = [
        Sheet("B2B", [Column("k", "K")], [{"k": "a"}]),
        Sheet("B2CL", [Column("k", "K")], [{"k": "b"}]),
        Sheet("B2CS", [Column("k", "K")], [{"k": "c"}]),
        Sheet("Export", [Column("k", "K")], [{"k": "d"}]),
        Sheet("HSN", [Column("k", "K")], [{"k": "e"}]),
    ]
    raw = to_xlsx(sheets)
    wb = load_workbook(io.BytesIO(raw))
    assert wb.sheetnames == ["B2B", "B2CL", "B2CS", "Export", "HSN"]


def test_xlsx_rejects_float_money() -> None:
    sheet = Sheet(
        name="X",
        columns=[Column("amount", "Amount", "money")],
        rows=[{"amount": 9.99}],
    )
    with pytest.raises(TypeError, match="float"):
        to_xlsx([sheet])


def test_media_types_match_rfc() -> None:
    assert CSV_MEDIA_TYPE.startswith("text/csv")
    assert XLSX_MEDIA_TYPE == ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def test_content_disposition_quotes_filename() -> None:
    assert content_disposition("invoices-2026-05-11.csv") == (
        'attachment; filename="invoices-2026-05-11.csv"'
    )


# ──────────────────────────────────────────────────────────────────────
# INJ-01 / RPT-03: formula-injection neutralisation (OWASP CWE-1236)
# ──────────────────────────────────────────────────────────────────────


_FORMULA_TRIGGERS = ["=cmd", "+cmd", "-cmd", "@SUM(A1)", "\tcmd", "\rcmd"]


def test_csv_sanitizes_formula_injection_triggers() -> None:
    """Values starting with = + - @ TAB CR must be prefixed with ' in CSV output.

    OWASP CWE-1236: without the prefix, a cell like ``=SYSTEM("cmd")`` is
    executed as a formula when the CSV is opened in Excel/LibreOffice.
    """
    columns = [Column("name", "Name", "text")]
    for trigger in _FORMULA_TRIGGERS:
        out = to_csv([{"name": trigger}], columns).decode("utf-8")
        expected = f"'{trigger}"
        assert expected in out, (
            f"Expected CSV cell to start with ' for trigger {trigger!r}; got:\n{out}"
        )


def test_csv_safe_values_unchanged() -> None:
    """Normal strings that don't start with formula triggers are not modified."""
    columns = [Column("name", "Name", "text")]
    for safe in ("Anjali Saree Centre", "123", "Total ₹1050", ""):
        out = to_csv([{"name": safe}], columns).decode("utf-8")
        # The value should appear verbatim (no leading apostrophe added).
        lines = out.splitlines()
        assert len(lines) >= 2
        # The data row should not start with a ' for safe values.
        cell_val = lines[1]
        if safe:
            assert not cell_val.startswith("'"), (
                f"Safe value {safe!r} should not be prefixed; got {cell_val!r}"
            )


def test_xlsx_sanitizes_formula_injection_triggers() -> None:
    """Text cells starting with formula triggers must be prefixed with ' in XLSX.

    After a round-trip through openpyxl, the cell value should start with
    an apostrophe so Excel treats it as text.  Note: the xlsx format
    normalises bare ``\\r`` to ``\\n`` internally, so we check
    ``startswith("'")`` rather than exact equality to avoid false failures
    on that platform-neutral normalisation.
    """
    for trigger in _FORMULA_TRIGGERS:
        sheet = Sheet(
            name="InjTest",
            columns=[Column("val", "Val", "text")],
            rows=[{"val": trigger}],
        )
        raw = to_xlsx([sheet])
        wb = load_workbook(io.BytesIO(raw))
        cell_val = wb["InjTest"]["A2"].value
        assert isinstance(cell_val, str) and cell_val.startswith("'"), (
            f"XLSX cell for trigger {trigger!r} should start with ', got {cell_val!r}"
        )


def test_xlsx_safe_text_values_unchanged() -> None:
    """Normal text values are not prefixed with ' in XLSX cells."""
    for safe in ("Anjali Saree Centre", "Total"):
        sheet = Sheet(
            name="Safe",
            columns=[Column("val", "Val", "text")],
            rows=[{"val": safe}],
        )
        raw = to_xlsx([sheet])
        wb = load_workbook(io.BytesIO(raw))
        cell_val = wb["Safe"]["A2"].value
        assert cell_val == safe, f"Safe value {safe!r} should be unchanged, got {cell_val!r}"
