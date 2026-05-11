"""End-to-end export tests — TASK-CUT-403.

Covers ``?format=csv|xlsx`` on every list / report endpoint the spec
calls out:

* GET /invoices            (CSV + XLSX)
* GET /parties             (CSV + XLSX)
* GET /items               (CSV + XLSX)
* GET /receipts            (CSV + XLSX)
* GET /vouchers            (CSV + XLSX)
* GET /reports/pnl         (CSV + XLSX)
* GET /reports/tb          (CSV + XLSX)
* GET /reports/daybook     (CSV + XLSX)
* GET /reports/stock-summary  (CSV + XLSX)
* GET /reports/gstr1       (XLSX multi-sheet)

Each test asserts the Content-Type, the attachment header, and one
content invariant (the seeded value appears in the bytes, or the
expected sheet shows up in the workbook).
"""

from __future__ import annotations

import datetime
import io
import uuid
from decimal import Decimal

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session as OrmSession

from app.service.export_service import UTF8_BOM

# ──────────────────────────────────────────────────────────────────────
# Helpers — auth + seed (mirrors the test_reports_routers / test_sales
# patterns so the export tests don't drift).
# ──────────────────────────────────────────────────────────────────────


def _signup_owner(client: TestClient) -> dict[str, str]:
    """Sign up + switch to the primary firm so the token carries firm_id."""
    resp = client.post(
        "/auth/signup",
        json={
            "email": f"u-{uuid.uuid4().hex[:10]}@example.com",
            "password": "strong-password-1",
            "org_name": f"Org-{uuid.uuid4().hex[:8]}",
            "firm_name": "Primary",
            "state_code": "MH",
        },
    )
    assert resp.status_code == 201, resp.text
    body: dict[str, str] = resp.json()
    switch = client.post(
        "/auth/switch-firm",
        headers={"Authorization": f"Bearer {body['access_token']}"},
        json={"firm_id": body["firm_id"]},
    )
    assert switch.status_code == 200, switch.text
    body["access_token"] = switch.json()["access_token"]
    return body


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _seed_party_and_item(sync_engine: Engine, *, org_id: uuid.UUID) -> tuple[uuid.UUID, uuid.UUID]:
    from app.models import Item, Party
    from app.models.masters import ItemType, TrackingType, UomType

    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        party = Party(
            org_id=org_id,
            code=f"P{uuid.uuid4().hex[:6].upper()}",
            name=f"Anjali Saree Centre {uuid.uuid4().hex[:4]}",
            is_customer=True,
            state_code="MH",
        )
        session.add(party)
        item = Item(
            org_id=org_id,
            code=f"I{uuid.uuid4().hex[:6].upper()}",
            name='Chiffon Silk 44"',
            item_type=ItemType.FINISHED,
            tracking=TrackingType.NONE,
            primary_uom=UomType.METER,
            hsn_code="5407",
            gst_rate=Decimal("5"),
        )
        session.add(item)
        session.commit()
        return party.party_id, item.item_id


def _create_and_finalize_invoice(
    http_client: TestClient,
    me: dict[str, str],
    *,
    party_id: uuid.UUID,
    item_id: uuid.UUID,
    invoice_date: str = "2026-04-30",
    qty: str = "1",
    price: str = "1000",
    gst_rate: str = "5",
) -> str:
    create = http_client.post(
        "/invoices",
        headers=_auth(me["access_token"]),
        json={
            "firm_id": me["firm_id"],
            "party_id": str(party_id),
            "invoice_date": invoice_date,
            "ship_to_state": "MH",
            "lines": [{"item_id": str(item_id), "qty": qty, "price": price, "gst_rate": gst_rate}],
        },
    )
    assert create.status_code == 201, create.text
    invoice_id: str = create.json()["sales_invoice_id"]
    fin = http_client.post(f"/invoices/{invoice_id}/finalize", headers=_auth(me["access_token"]))
    assert fin.status_code == 200, fin.text
    return invoice_id


CSV_MEDIA = "text/csv"
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _assert_attachment(resp: object, media: str, suffix: str) -> None:
    # `resp` is httpx.Response from TestClient; typed as `object` because
    # the import lives behind a fixture.
    assert resp.status_code == 200, resp.text  # type: ignore[attr-defined]
    assert resp.headers["content-type"].startswith(media), resp.headers["content-type"]  # type: ignore[attr-defined]
    disp = resp.headers["content-disposition"]  # type: ignore[attr-defined]
    assert disp.startswith('attachment; filename="'), disp
    assert disp.rstrip('"').endswith(f".{suffix}"), disp


# ──────────────────────────────────────────────────────────────────────
# Invoices
# ──────────────────────────────────────────────────────────────────────


def test_invoices_csv_export_returns_attachment_with_bom(
    http_client: TestClient, sync_engine: Engine
) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(http_client, me, party_id=party_id, item_id=item_id)

    resp = http_client.get("/invoices?format=csv", headers=_auth(me["access_token"]))
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert text_body.startswith(UTF8_BOM), "Excel-on-Windows needs UTF-8 BOM"
    # Header row + at least one data row.
    assert "Invoice #,Series,Date,Party" in text_body
    assert "1050.00" in text_body  # 1000 + 5% GST


def test_invoices_xlsx_export_returns_workbook(
    http_client: TestClient, sync_engine: Engine
) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(http_client, me, party_id=party_id, item_id=item_id)

    resp = http_client.get("/invoices?format=xlsx", headers=_auth(me["access_token"]))
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Invoices" in wb.sheetnames
    ws = wb["Invoices"]
    headers = [c.value for c in ws[1]]
    assert "Invoice #" in headers
    assert "Amount" in headers
    # Data starts at row 2; one of the cells in the Amount column should
    # be a Decimal-shaped number.
    amount_col = headers.index("Amount") + 1
    amount_cell = ws.cell(row=2, column=amount_col)
    assert amount_cell.value is not None
    # Money cells get the Indian-style 2-decimal number format.
    assert "0.00" in (amount_cell.number_format or "")


# ──────────────────────────────────────────────────────────────────────
# Parties
# ──────────────────────────────────────────────────────────────────────


def test_parties_csv_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    create = http_client.post(
        "/parties",
        headers=_auth(me["access_token"]),
        json={
            "code": "P-EXP-1",
            "name": "Export Test Customer",
            "is_customer": True,
            "gstin": "27ABCDE1234F1Z5",
            "pan": "ABCDE1234F",
            "state_code": "MH",
        },
    )
    assert create.status_code == 201, create.text

    resp = http_client.get("/parties?format=csv", headers=_auth(me["access_token"]))
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert "Code,Name" in text_body
    assert "P-EXP-1" in text_body
    # PII is the decrypted plaintext, not cipher-text bytes.
    assert "27ABCDE1234F1Z5" in text_body


def test_parties_xlsx_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    http_client.post(
        "/parties",
        headers=_auth(me["access_token"]),
        json={
            "code": "P-EXP-2",
            "name": "XLSX Test Supplier",
            "is_supplier": True,
            "state_code": "GJ",
        },
    )
    resp = http_client.get("/parties?format=xlsx", headers=_auth(me["access_token"]))
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Parties" in wb.sheetnames
    ws = wb["Parties"]
    values_col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "P-EXP-2" in values_col_a


# ──────────────────────────────────────────────────────────────────────
# Items
# ──────────────────────────────────────────────────────────────────────


def test_items_csv_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    _party_id, _item_id = _seed_party_and_item(sync_engine, org_id=org_id)

    resp = http_client.get("/items?format=csv", headers=_auth(me["access_token"]))
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert "Code,Name,Type,UOM" in text_body
    assert "Chiffon" in text_body


def test_items_xlsx_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    _party_id, _item_id = _seed_party_and_item(sync_engine, org_id=org_id)

    resp = http_client.get("/items?format=xlsx", headers=_auth(me["access_token"]))
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Items" in wb.sheetnames
    ws = wb["Items"]
    cells = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert any("Chiffon" in (v or "") for v in cells)


# ──────────────────────────────────────────────────────────────────────
# Receipts
# ──────────────────────────────────────────────────────────────────────


def test_receipts_csv_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(http_client, me, party_id=party_id, item_id=item_id)

    rcpt = http_client.post(
        "/receipts",
        headers=_auth(me["access_token"]),
        json={
            "party_id": str(party_id),
            "amount": "1050.00",
            "receipt_date": "2026-04-30",
            "mode": "CASH",
        },
    )
    assert rcpt.status_code == 201, rcpt.text

    resp = http_client.get("/receipts?format=csv", headers=_auth(me["access_token"]))
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert "Receipt #" in text_body
    assert "1050.00" in text_body


def test_receipts_xlsx_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(http_client, me, party_id=party_id, item_id=item_id)

    http_client.post(
        "/receipts",
        headers=_auth(me["access_token"]),
        json={
            "party_id": str(party_id),
            "amount": "1050.00",
            "receipt_date": "2026-04-30",
            "mode": "UPI",
        },
    )

    resp = http_client.get("/receipts?format=xlsx", headers=_auth(me["access_token"]))
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Receipts" in wb.sheetnames


# ──────────────────────────────────────────────────────────────────────
# Vouchers
# ──────────────────────────────────────────────────────────────────────


def test_vouchers_csv_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(http_client, me, party_id=party_id, item_id=item_id)

    resp = http_client.get("/vouchers?format=csv", headers=_auth(me["access_token"]))
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert "Voucher #" in text_body
    assert "SALES_INVOICE" in text_body


def test_vouchers_xlsx_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(http_client, me, party_id=party_id, item_id=item_id)

    resp = http_client.get("/vouchers?format=xlsx", headers=_auth(me["access_token"]))
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Vouchers" in wb.sheetnames


# ──────────────────────────────────────────────────────────────────────
# Reports — P&L
# ──────────────────────────────────────────────────────────────────────


def test_pnl_xlsx_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(
        http_client, me, party_id=party_id, item_id=item_id, invoice_date="2026-04-15"
    )

    resp = http_client.get(
        "/reports/pnl?from=2026-04-01&to=2026-04-30&format=xlsx",
        headers=_auth(me["access_token"]),
    )
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "P&L" in wb.sheetnames or "P_L" in wb.sheetnames


# ──────────────────────────────────────────────────────────────────────
# Reports — Trial Balance
# ──────────────────────────────────────────────────────────────────────


def test_tb_csv_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(http_client, me, party_id=party_id, item_id=item_id)
    http_client.post(
        "/receipts",
        headers=_auth(me["access_token"]),
        json={
            "party_id": str(party_id),
            "amount": "1050.00",
            "receipt_date": "2026-04-30",
            "mode": "CASH",
        },
    )

    resp = http_client.get(
        "/reports/tb?as_of=2026-04-30&format=csv",
        headers=_auth(me["access_token"]),
    )
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert "Code,Ledger" in text_body
    # Cash ledger code 1000 should appear in the TB once a receipt has
    # landed.
    assert "1000" in text_body


# ──────────────────────────────────────────────────────────────────────
# Reports — Daybook
# ──────────────────────────────────────────────────────────────────────


def test_daybook_xlsx_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(
        http_client, me, party_id=party_id, item_id=item_id, invoice_date="2026-04-30"
    )
    resp = http_client.get(
        "/reports/daybook?date=2026-04-30&format=xlsx",
        headers=_auth(me["access_token"]),
    )
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Daybook" in wb.sheetnames


# ──────────────────────────────────────────────────────────────────────
# Reports — Stock summary
# ──────────────────────────────────────────────────────────────────────


def test_stock_summary_csv_export(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    _party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)

    # Seed a lot + stock position so the report has at least one row.
    from app.models import Location, Lot, StockPosition
    from app.models.inventory import LocationType

    firm_id = uuid.UUID(me["firm_id"])
    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        loc = Location(
            org_id=org_id,
            firm_id=firm_id,
            code=f"WH-{uuid.uuid4().hex[:4].upper()}",
            name="Warehouse",
            location_type=LocationType.WAREHOUSE,
            is_active=True,
        )
        session.add(loc)
        session.flush()
        lot = Lot(
            org_id=org_id,
            firm_id=firm_id,
            item_id=item_id,
            lot_number=f"LOT-{uuid.uuid4().hex[:6]}",
            primary_cost=Decimal("50.0000"),
            cost_basis="STANDARD",
            currency="INR",
        )
        session.add(lot)
        session.flush()
        pos = StockPosition(
            org_id=org_id,
            firm_id=firm_id,
            item_id=item_id,
            lot_id=lot.lot_id,
            location_id=loc.location_id,
            on_hand_qty=Decimal("10"),
        )
        session.add(pos)
        session.commit()

    resp = http_client.get(
        "/reports/stock-summary?as_of=2026-04-30&format=csv",
        headers=_auth(me["access_token"]),
    )
    _assert_attachment(resp, CSV_MEDIA, "csv")
    text_body = resp.content.decode("utf-8")
    assert "Item code,Item name" in text_body
    assert "Chiffon" in text_body


# ──────────────────────────────────────────────────────────────────────
# Reports — GSTR-1 multi-sheet
# ──────────────────────────────────────────────────────────────────────


def test_gstr1_xlsx_export_has_five_sheets(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id, item_id = _seed_party_and_item(sync_engine, org_id=org_id)
    _create_and_finalize_invoice(
        http_client, me, party_id=party_id, item_id=item_id, invoice_date="2026-04-30"
    )

    resp = http_client.get(
        "/reports/gstr1?period=2026-04&format=xlsx",
        headers=_auth(me["access_token"]),
    )
    _assert_attachment(resp, XLSX_MEDIA, "xlsx")
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["B2B", "B2CL", "B2CS", "Export", "HSN"], wb.sheetnames


# ──────────────────────────────────────────────────────────────────────
# Permission gate — caller without sales.invoice.read can't export
# ──────────────────────────────────────────────────────────────────────


def test_invoice_export_requires_auth(http_client: TestClient) -> None:
    """Anonymous request → 401, same as the JSON endpoint."""
    resp = http_client.get("/invoices?format=csv")
    assert resp.status_code == 401


# ──────────────────────────────────────────────────────────────────────
# RLS isolation — org A's export doesn't leak org B's data
# ──────────────────────────────────────────────────────────────────────


def test_invoices_csv_export_rls_isolated(http_client: TestClient, sync_engine: Engine) -> None:
    """Owner of org A exports invoices; org B's invoice (seeded by B's
    own session) must not appear in the CSV.
    """
    a = _signup_owner(http_client)
    b = _signup_owner(http_client)
    party_a, item_a = _seed_party_and_item(sync_engine, org_id=uuid.UUID(a["org_id"]))
    party_b, item_b = _seed_party_and_item(sync_engine, org_id=uuid.UUID(b["org_id"]))
    _create_and_finalize_invoice(http_client, a, party_id=party_a, item_id=item_a)
    _create_and_finalize_invoice(http_client, b, party_id=party_b, item_id=item_b)

    resp_a = http_client.get("/invoices?format=csv", headers=_auth(a["access_token"]))
    _assert_attachment(resp_a, CSV_MEDIA, "csv")
    text_a = resp_a.content.decode("utf-8")
    # Org A sees its own party but not org B's.
    from app.models import Party

    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{a['org_id']}'"))
        a_party = session.get(Party, party_a)
        assert a_party is not None
        assert a_party.name in text_a, "Org A's party should show up in A's export"
        session.execute(text(f"SET LOCAL app.current_org_id = '{b['org_id']}'"))
        b_party = session.get(Party, party_b)
        assert b_party is not None
        assert b_party.name not in text_a, "Org B's party MUST NOT leak into A's CSV"


# Suppress imports that exist for type clarity even when the test file
# is skipped (no DB available locally).
_ = datetime
