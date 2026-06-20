"""TASK-CUT-302: ``GET /reports/gstr1?period=YYYY-MM`` integration tests.

Buckets exercised:
  - B2B (party with GSTIN set)
  - B2CL (B2C inter-state > ₹2.5L)
  - B2CS (B2C aggregated by state + rate)
  - Export (party.is_export / .is_sez)
  - HSN summary (aggregation across all invoice lines)
"""

from __future__ import annotations

import io
import uuid
from decimal import Decimal

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session as OrmSession

from tests.test_reports_routers import (
    _auth,
    _create_and_finalize_invoice,
    _signup_owner,
)


def _create_invoice_with_ship_to(
    http_client: TestClient,
    me: dict[str, str],
    *,
    party_id: uuid.UUID,
    item_id: uuid.UUID,
    invoice_date: str,
    ship_to_state: str,
    qty: str = "1",
    price: str = "1000",
    gst_rate: str = "5",
) -> str:
    """Like ``_create_and_finalize_invoice`` but with an explicit ship_to_state.

    The shared helper hard-codes ship_to_state to MH, which means every
    invoice resolves to the seller's state (intra-state). GSTR-1 tests
    need to drive cross-state PoS by overriding it.
    """
    create = http_client.post(
        "/invoices",
        headers=_auth(me["access_token"]),
        json={
            "firm_id": me["firm_id"],
            "party_id": str(party_id),
            "invoice_date": invoice_date,
            "ship_to_state": ship_to_state,
            "lines": [{"item_id": str(item_id), "qty": qty, "price": price, "gst_rate": gst_rate}],
        },
    )
    assert create.status_code == 201, create.text
    invoice_id: str = create.json()["sales_invoice_id"]
    fin = http_client.post(f"/invoices/{invoice_id}/finalize", headers=_auth(me["access_token"]))
    assert fin.status_code == 200, fin.text
    return invoice_id


def _seed_b2b_party(sync_engine: Engine, *, org_id: uuid.UUID, state_code: str = "MH") -> uuid.UUID:
    """Seed a customer party with a GSTIN set so the PoS engine
    classifies the buyer as REGISTERED."""
    from app.models import Party

    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        party = Party(
            org_id=org_id,
            code=f"B2B{uuid.uuid4().hex[:6].upper()}",
            name=f"B2B {uuid.uuid4().hex[:4]}",
            is_customer=True,
            state_code=state_code,
            gstin=b"\x27" * 15,  # 15 bytes of dummy ciphertext (GSTIN is encrypted in DB)
        )
        session.add(party)
        session.commit()
        return party.party_id


def _seed_b2c_party(sync_engine: Engine, *, org_id: uuid.UUID, state_code: str = "GJ") -> uuid.UUID:
    """Seed a customer party WITHOUT a GSTIN — CONSUMER for PoS engine."""
    from app.models import Party

    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        party = Party(
            org_id=org_id,
            code=f"B2C{uuid.uuid4().hex[:6].upper()}",
            name=f"B2C {uuid.uuid4().hex[:4]}",
            is_customer=True,
            state_code=state_code,
        )
        session.add(party)
        session.commit()
        return party.party_id


def _seed_item(sync_engine: Engine, *, org_id: uuid.UUID, hsn_code: str = "5208") -> uuid.UUID:
    """Seed a finished item with an HSN code."""
    from app.models import Item
    from app.models.masters import ItemType, TrackingType, UomType

    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        item = Item(
            org_id=org_id,
            code=f"I{uuid.uuid4().hex[:6].upper()}",
            name="Chiffon",
            item_type=ItemType.FINISHED,
            tracking=TrackingType.NONE,
            primary_uom=UomType.METER,
            hsn_code=hsn_code,
        )
        session.add(item)
        session.commit()
        return item.item_id


def test_gstr1_empty_for_fresh_firm(http_client: TestClient, sync_engine: Engine) -> None:
    me = _signup_owner(http_client)
    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["period"] == "2026-04"
    assert body["from_date"] == "2026-04-01"
    assert body["to_date"] == "2026-04-30"
    assert body["b2b"] == []
    assert body["b2cl"] == []
    assert body["b2cs"] == []
    assert body["export"] == []
    assert body["hsn"] == []


def _seed_b2b_party_with_real_gstin(
    sync_engine: Engine,
    *,
    org_id: uuid.UUID,
    gstin: str,
    state_code: str = "MH",
) -> uuid.UUID:
    """Seed a B2B party whose GSTIN is REAL plaintext run through the
    production encryption path (so the DB row holds AES-GCM ciphertext).

    Used by the B2 regression test that proves GSTR-1 reports plaintext,
    not `hex(ciphertext)`.
    """
    from app.models import Party
    from app.utils.crypto import encrypt_pii, get_org_dek

    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        dek = get_org_dek(session, org_id=org_id)
        party = Party(
            org_id=org_id,
            code=f"B2B{uuid.uuid4().hex[:6].upper()}",
            name=f"B2B {uuid.uuid4().hex[:4]}",
            is_customer=True,
            state_code=state_code,
            gstin=encrypt_pii(gstin, dek=dek, org_id=org_id),
        )
        session.add(party)
        session.commit()
        return party.party_id


def test_gstr1_b2b_returns_plaintext_gstin_not_ciphertext_hex(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """B2 fix: the B2B bucket's `gstin` field must be the *plaintext*
    GSTIN as filed to GSTN, not `hex(ciphertext)`.

    Before the fix, `compute_gstr1` rendered `r.party_gstin.hex()` —
    which is hex of an AES-GCM ciphertext that's per-encryption unique.
    That breaks both filing (GSTN rejects non-15-char values) and B2B
    aggregation across parties that share a plaintext GSTIN (e.g. multi-
    branch customers).
    """
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_gstin = "27ABCDE1234F1Z5"  # realistic MH GSTIN — 15 chars
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin=party_gstin, state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["b2b"]) == 1
    inv = body["b2b"][0]
    # The whole point of B2: GSTR-1 must surface the plaintext, never
    # the ciphertext hex. Anything other than the exact filed GSTIN is
    # a regression — GSTN would reject it, and downstream B2B aggregation
    # (multiple invoices to the same registered party) would split rows.
    assert inv["gstin"] == party_gstin, (
        f"GSTR-1 must return plaintext GSTIN {party_gstin!r}, got {inv['gstin']!r} — "
        f"reports_service is still emitting hex(ciphertext) instead of decrypting."
    )


def test_gstr1_b2b_bucket_for_registered_party(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """A party with a GSTIN set → invoice lands in B2B bucket."""
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id = _seed_b2b_party(sync_engine, org_id=org_id, state_code="MH")
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["b2b"]) == 1
    inv = body["b2b"][0]
    assert inv["party_id"] == str(party_id)
    assert Decimal(inv["taxable_value"]) == Decimal("1000.00")
    assert Decimal(inv["invoice_value"]) == Decimal("1050.00")
    # MH→MH intra-state → CGST+SGST split, no IGST.
    assert Decimal(inv["cgst"]) == Decimal("25.00")
    assert Decimal(inv["sgst"]) == Decimal("25.00")
    assert Decimal(inv["igst"]) == Decimal("0")
    # HSN summary aggregates the single line.
    assert len(body["hsn"]) == 1
    hsn_row = body["hsn"][0]
    assert hsn_row["hsn_code"] == "5208"
    assert Decimal(hsn_row["total_qty"]) == Decimal("2")
    assert Decimal(hsn_row["taxable_value"]) == Decimal("1000.00")
    assert Decimal(hsn_row["cgst"]) == Decimal("25.00")
    assert Decimal(hsn_row["sgst"]) == Decimal("25.00")


def test_gstr1_b2cl_for_inter_state_above_threshold(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """B2C inter-state invoice > ₹2.5L → B2CL bucket."""
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    # Seller is MH (signup default); customer in GJ (inter-state), no GSTIN.
    party_id = _seed_b2c_party(sync_engine, org_id=org_id, state_code="GJ")
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    # Subtotal 300_000 → triggers > ₹2.5L threshold.
    _create_invoice_with_ship_to(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        ship_to_state="GJ",
        qty="1",
        price="300000",
        gst_rate="18",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["b2cl"]) == 1
    row = body["b2cl"][0]
    assert Decimal(row["taxable_value"]) == Decimal("300000.00")
    assert Decimal(row["igst"]) == Decimal("54000.00")
    assert Decimal(row["cgst"]) == Decimal("0")
    assert body["b2cs"] == []
    assert body["b2b"] == []


def test_gstr1_b2cs_aggregates_small_invoices_by_state_rate(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """Two small B2C inter-state invoices, same (state, rate) → one
    aggregated row in B2CS; two intra-state B2C invoices → another row."""
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    # Inter-state customer.
    party_inter = _seed_b2c_party(sync_engine, org_id=org_id, state_code="GJ")
    # Intra-state customer (MH).
    party_intra = _seed_b2c_party(sync_engine, org_id=org_id, state_code="MH")
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")

    # Two small inter-state invoices ₹1000 each at 5%.
    for _ in range(2):
        _create_invoice_with_ship_to(
            http_client,
            me,
            party_id=party_inter,
            item_id=item_id,
            invoice_date="2026-04-15",
            ship_to_state="GJ",
            qty="1",
            price="1000",
            gst_rate="5",
        )
    # One intra-state ₹2000 at 5%.
    _create_invoice_with_ship_to(
        http_client,
        me,
        party_id=party_intra,
        item_id=item_id,
        invoice_date="2026-04-20",
        ship_to_state="MH",
        qty="1",
        price="2000",
        gst_rate="5",
    )

    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    # B2CS has two rows: (GJ, 5%) and (MH, 5%).
    b2cs = {(r["place_of_supply_state"], Decimal(r["gst_rate"])): r for r in body["b2cs"]}
    assert len(b2cs) == 2
    gj = b2cs[("GJ", Decimal("5"))]
    assert Decimal(gj["taxable_value"]) == Decimal("2000.00")
    assert Decimal(gj["igst"]) == Decimal("100.00")
    assert gj["invoice_count"] == 2
    mh = b2cs[("MH", Decimal("5"))]
    assert Decimal(mh["taxable_value"]) == Decimal("2000.00")
    assert Decimal(mh["cgst"]) == Decimal("50.00")
    assert Decimal(mh["sgst"]) == Decimal("50.00")
    assert Decimal(mh["igst"]) == Decimal("0")


def test_gstr1_export_bucket_for_export_party(http_client: TestClient, sync_engine: Engine) -> None:
    """Party with is_export=True → invoice lands in export bucket."""
    from app.models import Party

    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        party = Party(
            org_id=org_id,
            code=f"EXP{uuid.uuid4().hex[:6].upper()}",
            name="Export Customer",
            is_customer=True,
            is_export=True,
        )
        session.add(party)
        session.commit()
        party_id = party.party_id
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="1",
        price="1000",
        gst_rate="0",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["export"]) == 1
    row = body["export"][0]
    assert row["party_id"] == str(party_id)
    assert Decimal(row["taxable_value"]) == Decimal("1000.00")
    # Zero-rated → no GST.
    assert Decimal(row["igst"]) == Decimal("0")
    assert body["b2b"] == []
    assert body["b2cs"] == []


def test_gstr1_rls_isolated_across_orgs(http_client: TestClient, sync_engine: Engine) -> None:
    a = _signup_owner(http_client)
    b = _signup_owner(http_client)
    org_a = uuid.UUID(a["org_id"])
    party_id = _seed_b2b_party(sync_engine, org_id=org_a, state_code="MH")
    item_id = _seed_item(sync_engine, org_id=org_a, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        a,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="1",
        price="1000",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(b["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["b2b"] == [], "B saw A's B2B invoice — RLS leak"
    assert body["hsn"] == []


def test_gstr1_requires_report_view_permission(
    http_client: TestClient, sync_engine: Engine
) -> None:
    from app.models import AppUser, Role
    from app.service import identity_service, rbac_service

    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        sales_role = session.execute(
            select(Role).where(Role.org_id == org_id, Role.code == "SALESPERSON")
        ).scalar_one()
        sales_user = identity_service.register_user(
            session,
            email=f"sales-{uuid.uuid4().hex[:6]}@example.com",
            password="strong-password-1",
            org_id=org_id,
        )
        rbac_service.assign_role(
            session,
            user_id=sales_user.user_id,
            role_id=sales_role.role_id,
            firm_id=uuid.UUID(me["firm_id"]),
            org_id=org_id,
        )
        sales_user_id = sales_user.user_id
        session.commit()
    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        sales_user = session.execute(
            select(AppUser).where(AppUser.user_id == sales_user_id)
        ).scalar_one()
        pair = identity_service.issue_tokens(
            session, user=sales_user, firm_id=uuid.UUID(me["firm_id"])
        )
        session.commit()
    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(pair.access_token),
    )
    assert resp.status_code == 403, resp.text
    assert resp.json()["code"] == "PERMISSION_DENIED"


# ──────────────────────────────────────────────────────────────────────
# TASK-TR-Q05a — XLSX export column-name mismatch regressions
#
# `export_builders.GSTR1_*_COLUMNS` declared keys like ``party_gstin`` /
# ``invoice_number`` / ``total_quantity`` / ``cgst_amount`` etc, but the
# row dataclasses (`_Gstr1InvoiceRow`, `_Gstr1HsnRow`, `_Gstr1B2csRow`)
# expose ``gstin`` / ``number`` / ``total_qty`` / ``cgst`` etc. The
# `_as_dict(row, columns)` helper does ``getattr(row, c.key, None)`` so
# the cells silently rendered empty. JSON API was unaffected (router
# maps dataclasses to pydantic models with the short names).
#
# These tests open the XLSX bytes with openpyxl and assert each affected
# cell carries the *seeded* value, not None.
# ──────────────────────────────────────────────────────────────────────


def _header_index(ws: object, header: str) -> int:
    """1-based column index for the given header text, or fail loudly."""
    headers = [c.value for c in ws[1]]  # type: ignore[index]
    assert header in headers, f"{header!r} missing from sheet headers {headers!r}"
    return headers.index(header) + 1


def test_gstr1_xlsx_b2b_sheet_contains_party_gstin(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """B2B sheet's GSTIN column must carry the plaintext GSTIN, not blank.

    Pre-fix: `Column("party_gstin", ...)` mismatched `_Gstr1InvoiceRow.gstin`,
    so every B2B row's GSTIN cell was empty in the exported workbook.
    """
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_gstin = "27ABCDE1234F1Z5"
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin=party_gstin, state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04&format=xlsx",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    assert "B2B" in wb.sheetnames, wb.sheetnames
    ws = wb["B2B"]
    gstin_col = _header_index(ws, "GSTIN")
    gstin_cell = ws.cell(row=2, column=gstin_col).value
    assert gstin_cell == party_gstin, (
        f"B2B sheet GSTIN cell must be plaintext GSTIN {party_gstin!r}, "
        f"got {gstin_cell!r} — column-key mismatch is back."
    )


def test_gstr1_xlsx_b2b_sheet_contains_invoice_number(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """B2B sheet's Invoice # column must carry the source invoice number.

    Pre-fix: `Column("invoice_number", ...)` mismatched
    `_Gstr1InvoiceRow.number`, so every Invoice # cell was empty.
    """
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin="27ABCDE1234F1Z5", state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )

    json_resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(me["access_token"]),
    )
    assert json_resp.status_code == 200, json_resp.text
    expected_number = json_resp.json()["b2b"][0]["number"]

    resp = http_client.get(
        "/reports/gstr1?period=2026-04&format=xlsx",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["B2B"]
    inv_col = _header_index(ws, "Invoice #")
    inv_cell = ws.cell(row=2, column=inv_col).value
    assert inv_cell, f"Invoice # cell is empty: {inv_cell!r}"
    assert str(inv_cell) == str(expected_number), (
        f"Expected Invoice # {expected_number!r}, got {inv_cell!r}"
    )


def test_gstr1_xlsx_hsn_sheet_contains_total_quantity(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """HSN sheet's Qty column must carry the summed quantity.

    Pre-fix: `Column("total_quantity", ...)` mismatched
    `_Gstr1HsnRow.total_qty`, so every Qty cell was empty.
    """
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin="27ABCDE1234F1Z5", state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04&format=xlsx",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    assert "HSN" in wb.sheetnames, wb.sheetnames
    ws = wb["HSN"]
    qty_col = _header_index(ws, "Qty")
    qty_cell = ws.cell(row=2, column=qty_col).value
    assert qty_cell is not None, "HSN Qty cell is empty"
    assert Decimal(str(qty_cell)) == Decimal("2"), f"Expected total qty 2, got {qty_cell!r}"


def test_gstr1_xlsx_b2b_sheet_contains_tax_amounts(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """B2B sheet's CGST/SGST/IGST columns must carry the tax amounts.

    Pre-fix: `Column("cgst_amount", ...)` / `sgst_amount` / `igst_amount`
    all mismatched `_Gstr1InvoiceRow.cgst`/`.sgst`/`.igst`, so every tax
    cell was empty in the workbook. Same pattern as the GSTIN / number /
    total_qty bugs.
    """
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin="27ABCDE1234F1Z5", state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04&format=xlsx",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["B2B"]
    cgst_col = _header_index(ws, "CGST")
    sgst_col = _header_index(ws, "SGST")
    igst_col = _header_index(ws, "IGST")
    cgst_cell = ws.cell(row=2, column=cgst_col).value
    sgst_cell = ws.cell(row=2, column=sgst_col).value
    igst_cell = ws.cell(row=2, column=igst_col).value
    # MH→MH intra-state: CGST 2.5% + SGST 2.5% on ₹1000 = ₹25 each;
    # IGST is 0.
    assert cgst_cell is not None and Decimal(str(cgst_cell)) == Decimal("25.00"), cgst_cell
    assert sgst_cell is not None and Decimal(str(sgst_cell)) == Decimal("25.00"), sgst_cell
    assert igst_cell is not None and Decimal(str(igst_cell)) == Decimal("0"), igst_cell


def test_gstr1_csv_b2b_contains_party_gstin_and_number(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """CSV export of the B2B sheet must contain the plaintext GSTIN and
    invoice number — same column-key mismatch affected CSV. The CSV
    branch flattens the B2B sheet only.
    """
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_gstin = "27ABCDE1234F1Z5"
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin=party_gstin, state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04&format=csv",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    text_body = resp.content.decode("utf-8")
    assert party_gstin in text_body, (
        f"Plaintext GSTIN {party_gstin!r} missing from CSV body — column-key mismatch is back."
    )


# ──────────────────────────────────────────────────────────────────────
# RPT-02: GSTR-1 GSTIN masking without masters.party.read
# ──────────────────────────────────────────────────────────────────────


def test_gstr1_gstin_masked_when_can_view_pii_false(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """compute_gstr1(can_view_pii=False) must mask GSTIN to last-3 chars.

    Without masters.party.read the caller should see "***1Z5", not the
    full "27ABCDE1234F1Z5" — prevents PII leakage to lower-privilege
    accounting-report-view-only callers.
    """
    from app.service import reports_service
    from sqlalchemy.orm import Session as OrmSession

    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    firm_id = uuid.UUID(me["firm_id"])
    party_gstin = "27ABCDE1234F1Z5"
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin=party_gstin, state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )

    # Invoke service directly with can_view_pii=False (lower-privilege caller).
    sync_url = sync_engine.url
    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        result = reports_service.compute_gstr1(
            session,
            org_id=org_id,
            firm_id=firm_id,
            period="2026-04",
            can_view_pii=False,
        )

    assert len(result.b2b) == 1
    masked_gstin = result.b2b[0].gstin
    # Must be masked: last-3 chars visible, rest replaced with "*"
    assert masked_gstin is not None
    assert masked_gstin.endswith(party_gstin[-3:]), (
        f"Expected masked GSTIN ending with {party_gstin[-3:]!r}, got {masked_gstin!r}"
    )
    assert masked_gstin != party_gstin, (
        f"GSTIN was not masked: got full plaintext {masked_gstin!r}"
    )
    assert "*" in masked_gstin, (
        f"Expected '*' in masked GSTIN, got {masked_gstin!r}"
    )


def test_gstr1_gstin_full_when_can_view_pii_true(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """compute_gstr1(can_view_pii=True) must return the full plaintext GSTIN."""
    from app.service import reports_service
    from sqlalchemy.orm import Session as OrmSession

    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    firm_id = uuid.UUID(me["firm_id"])
    party_gstin = "27ABCDE1234F1Z5"
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin=party_gstin, state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )

    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        result = reports_service.compute_gstr1(
            session,
            org_id=org_id,
            firm_id=firm_id,
            period="2026-04",
            can_view_pii=True,
        )

    assert len(result.b2b) == 1
    assert result.b2b[0].gstin == party_gstin, (
        f"Expected full GSTIN {party_gstin!r} with can_view_pii=True, "
        f"got {result.b2b[0].gstin!r}"
    )


# ──────────────────────────────────────────────────────────────────────
# RPT-02 (cycle-2): GSTR-1 GSTIN reveal gated on masters.party.pii.read
# (not the broader masters.party.read) — HTTP-level router gate tests.
# ──────────────────────────────────────────────────────────────────────


def test_gstr1_gstin_revealed_for_user_with_pii_read_permission(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """Owner (has masters.party.pii.read) must receive the plaintext GSTIN
    in the HTTP response — the router gate should resolve can_view_pii=True.

    This is the HTTP-level companion to the service-level
    test_gstr1_gstin_full_when_can_view_pii_true — it proves the *router*
    checks the right permission string.
    """
    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_gstin = "27ABCDE1234F1Z5"
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin=party_gstin, state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )
    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(me["access_token"]),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["b2b"]) == 1
    assert body["b2b"][0]["gstin"] == party_gstin, (
        f"Owner with masters.party.pii.read must see full GSTIN {party_gstin!r}, "
        f"got {body['b2b'][0]['gstin']!r}"
    )


def test_gstr1_gstin_masked_for_user_without_pii_read_permission(
    http_client: TestClient, sync_engine: Engine
) -> None:
    """A report viewer WITHOUT masters.party.pii.read must receive a masked
    GSTIN in the HTTP response.

    This proves the router gate uses masters.party.pii.read specifically.
    The system ACCOUNTANT role is granted masters.party.pii.read (so real
    accountants keep PII access), so to exercise the masked path we mint a
    custom role that has accounting.report.view but NOT pii.read — without
    this test, reverting the gate to 'masters.party.read' (or dropping the
    pii.read grant) would slip through.
    """
    from app.models import AppUser
    from app.service import identity_service, rbac_service

    me = _signup_owner(http_client)
    org_id = uuid.UUID(me["org_id"])
    party_gstin = "27ABCDE1234F1Z5"
    party_id = _seed_b2b_party_with_real_gstin(
        sync_engine, org_id=org_id, gstin=party_gstin, state_code="MH"
    )
    item_id = _seed_item(sync_engine, org_id=org_id, hsn_code="5208")
    _create_and_finalize_invoice(
        http_client,
        me,
        party_id=party_id,
        item_id=item_id,
        invoice_date="2026-04-15",
        qty="2",
        price="500",
        gst_rate="5",
    )

    # Mint a custom role that can view reports + party names but explicitly
    # lacks masters.party.pii.read, then assign a fresh user to it.
    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        report_role = rbac_service.create_custom_role(
            session,
            org_id=org_id,
            code="REPORT_VIEWER_NO_PII",
            name="Report viewer without PII",
            permission_codes=["accounting.report.view", "masters.party.read"],
        )
        acct_user = identity_service.register_user(
            session,
            email=f"acct-pii-{uuid.uuid4().hex[:6]}@example.com",
            password="strong-password-1",
            org_id=org_id,
        )
        rbac_service.assign_role(
            session,
            user_id=acct_user.user_id,
            role_id=report_role.role_id,
            firm_id=uuid.UUID(me["firm_id"]),
            org_id=org_id,
        )
        acct_user_id = acct_user.user_id
        session.commit()

    with OrmSession(sync_engine, expire_on_commit=False) as session:
        session.execute(text(f"SET LOCAL app.current_org_id = '{org_id}'"))
        acct_user = session.execute(
            select(AppUser).where(AppUser.user_id == acct_user_id)
        ).scalar_one()
        pair = identity_service.issue_tokens(
            session, user=acct_user, firm_id=uuid.UUID(me["firm_id"])
        )
        session.commit()

    resp = http_client.get(
        "/reports/gstr1?period=2026-04",
        headers=_auth(pair.access_token),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["b2b"]) == 1
    gstin_in_response = body["b2b"][0]["gstin"]
    # Must be masked — ACCOUNTANT lacks masters.party.pii.read.
    assert gstin_in_response != party_gstin, (
        f"ACCOUNTANT must NOT see full GSTIN; got {gstin_in_response!r} "
        f"which equals the plaintext — router gate not checking pii.read"
    )
    assert gstin_in_response is not None and "*" in gstin_in_response, (
        f"Expected masked GSTIN (with '*'), got {gstin_in_response!r}"
    )
