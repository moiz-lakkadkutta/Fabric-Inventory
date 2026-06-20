"""TASK-CUT-402: VyaparExcelAdapter unit tests.

Pure-Python tests against the fixture xlsx — no DB required. The
adapter is stateless and side-effect-free, so these run in a few
milliseconds. Integration tests (the upload+approve flow) live in
``test_migrations_router.py``.

Fixtures:
    backend/tests/fixtures/vyapar-sample.xlsx — synthetic Vyapar
    parties export with 5 rows (2 customers with balance, 1 supplier
    with balance, 1 karigar with zero, 1 customer with no balance).
    Hand-crafted so the OB math sums to a balanced TB.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest

from app.service.migration import (
    IntermediateOpeningBalance,
    IntermediateParty,
    MigrationAdapter,
    VyaparExcelAdapter,
)
from app.service.migration.vyapar_adapter import _parse_decimal

_FIXTURE = Path(__file__).parent / "fixtures" / "vyapar-sample.xlsx"


def _bytes() -> bytes:
    return _FIXTURE.read_bytes()


def test_adapter_conforms_to_protocol() -> None:
    """Structural conformance — isinstance check against the Protocol."""
    assert isinstance(VyaparExcelAdapter(), MigrationAdapter)


def test_extract_parties_yields_expected_rows() -> None:
    """Parses parties, normalises kinds, derives codes from names."""
    adapter = VyaparExcelAdapter()
    parties = list(adapter.extract_parties(_bytes()))

    # 5 data rows in fixture, all have a name → 5 parties.
    assert len(parties) == 5
    for p in parties:
        assert isinstance(p, IntermediateParty)

    by_name = {p.name: p for p in parties}
    anjali = by_name["Anjali Saree Centre"]
    assert anjali.kinds == ("CUSTOMER",)
    assert anjali.gstin == "24AAACR5055K1Z5"
    assert anjali.state_code == "GJ"
    # Code is derived from the first 3 alpha tokens.
    assert anjali.code.startswith("ANJALI-")
    assert anjali.phone == "+91 9876543210"
    assert anjali.email == "anjali@example.com"

    silk = by_name["Surat Silk Mills"]
    assert silk.kinds == ("SUPPLIER",)
    assert silk.gstin == "24BBBCC5555K2Z9"

    karigar = by_name["Imran Karigar"]
    assert karigar.kinds == ("KARIGAR",)


def test_extract_opening_balances_balanced() -> None:
    """OB extraction is balanced (DR == CR) for the fixture file."""
    adapter = VyaparExcelAdapter()
    obs = list(adapter.extract_opening_balances(_bytes()))

    # 3 non-zero balances (2 customers + 1 supplier; karigar = 0, blank = skipped)
    assert len(obs) == 3
    for o in obs:
        assert isinstance(o, IntermediateOpeningBalance)
        # Each OB row references a party by source_id.
        assert o.party_source_id is not None

    dr_total = sum((o.amount for o in obs if o.side == "DR"), Decimal("0"))
    cr_total = sum((o.amount for o in obs if o.side == "CR"), Decimal("0"))

    # Expected from the fixture:
    #   Anjali (To Receive) — Customer — 15000 DR (Sundry Debtors)
    #   Devi   (To Receive) — Customer — 8500.50 DR (Sundry Debtors)
    #   Silk   (To Pay)     — Supplier — 23500.50 CR (Sundry Creditors)
    #   --------------------------------------------------
    #   DR 23500.50   ==   CR 23500.50  → balanced
    assert dr_total == Decimal("23500.50")
    assert cr_total == Decimal("23500.50")
    assert dr_total == cr_total  # the invariant


def test_opening_balance_kinds_match_party_kinds() -> None:
    """Customer balances hit SUNDRY_DEBTORS, supplier hits SUNDRY_CREDITORS."""
    adapter = VyaparExcelAdapter()
    parties = {p.source_id: p for p in adapter.extract_parties(_bytes())}
    obs = list(adapter.extract_opening_balances(_bytes()))

    for o in obs:
        party = parties[o.party_source_id]  # type: ignore[index]
        if "SUPPLIER" in party.kinds:
            assert o.ledger_kind == "SUNDRY_CREDITORS"
            assert o.side == "CR"
        else:
            assert o.ledger_kind == "SUNDRY_DEBTORS"
            assert o.side == "DR"


def test_validate_reports_balanced_count() -> None:
    """Validate returns matching party + OB counts; zero errors on a clean file."""
    report = VyaparExcelAdapter().validate(_bytes())

    assert report.total_parties == 5
    assert report.total_opening_balances == 3
    assert report.errors == 0
    # Tb_diff is left None by validate (commit step computes it).
    assert report.tb_diff is None
    # At least one info row about extraction.
    assert any(r.severity == "info" and r.code == "EXTRACTED" for r in report.rows)


def test_validate_reports_unparseable_balance_as_error() -> None:
    """A row with a non-numeric balance produces an error reconciliation row."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Parties")
    ws.append(["Party Name", "Opening Balance", "Type", "Balance Type"])
    ws.append(["Good Party", "1000", "Customer", "To Receive"])
    ws.append(["Bad Party", "not-a-number", "Customer", "To Receive"])
    out = BytesIO()
    wb.save(out)

    report = VyaparExcelAdapter().validate(out.getvalue())
    assert report.errors >= 1
    bad = [r for r in report.rows if r.code == "OPENING_BALANCE_UNPARSEABLE"]
    assert len(bad) == 1


def test_validate_empty_workbook_errors() -> None:
    """Workbook with no party sheet emits NO_PARTIES_FOUND."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Random"
    wb.active.append(["foo", "bar"])
    out = BytesIO()
    wb.save(out)

    report = VyaparExcelAdapter().validate(out.getvalue())
    assert report.total_parties == 0
    assert report.errors >= 1
    assert any(r.code == "NO_PARTIES_FOUND" for r in report.rows)


def test_parse_decimal_handles_indian_formatting() -> None:
    """Money parser strips ₹ and Indian commas without using float."""
    assert _parse_decimal("1234.50") == Decimal("1234.50")
    assert _parse_decimal("₹1,23,450.00") == Decimal("123450.00")
    assert _parse_decimal("₹ 1,000") == Decimal("1000")
    assert _parse_decimal(1234) == Decimal("1234")
    assert _parse_decimal(1234.5) == Decimal("1234.5")
    assert _parse_decimal(None) == Decimal("0")
    assert _parse_decimal("") == Decimal("0")
    assert _parse_decimal("-") == Decimal("0")
    with pytest.raises(ValueError):
        _parse_decimal("not-a-number")
    with pytest.raises(ValueError):
        _parse_decimal(True)


def test_invalid_gstin_warns_but_imports() -> None:
    """A bad GSTIN doesn't fail the row — it's emitted as a warning."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Parties")
    ws.append(["Party Name", "GSTIN", "Type", "Opening Balance", "Balance Type"])
    ws.append(["Mostly Good", "NOTAGSTIN", "Customer", "1000", "To Receive"])
    out = BytesIO()
    wb.save(out)

    adapter = VyaparExcelAdapter()
    report = adapter.validate(out.getvalue())
    assert report.warnings >= 1
    assert any(r.code == "GSTIN_FORMAT_INVALID" for r in report.rows)
    # Party still extracted, just without GSTIN.
    parties = list(adapter.extract_parties(out.getvalue()))
    assert len(parties) == 1
    assert parties[0].gstin is None


# ---------------------------------------------------------------------
# TASK-TR-D3-PREP: structural guards for the D3 spike runner.
#
# The runner at backend/scripts/spike_vyapar.py imports _COLUMN_MAP,
# _PARTY_SHEET_NAMES, _PARTY_TYPE_MAP, and VyaparExcelAdapter by name.
# These tests fail loudly if anyone renames those constants without
# updating the runner — without them, the spike silently breaks the
# day Moiz drops a real file.
# ---------------------------------------------------------------------


def test_d3_constants_present() -> None:
    """The D3 spike runner imports these four names by hard reference."""
    from app.service.migration import vyapar_adapter

    for name in ("_COLUMN_MAP", "_PARTY_SHEET_NAMES", "_PARTY_TYPE_MAP", "VyaparExcelAdapter"):
        assert hasattr(vyapar_adapter, name), (
            f"Vyapar adapter is missing {name!r}; the TR-D3-PREP spike runner "
            "imports it by name. Update backend/scripts/spike_vyapar.py if you "
            "rename it."
        )


def test_d3_column_map_values_target_known_fields() -> None:
    """Every _COLUMN_MAP target must be a real IntermediateParty field
    or one of the two adapter-internal sentinels (``_opening_balance``,
    ``_balance_type``, ``_party_type``). Catches a refactor that drops
    a field on the intermediate model but forgets to update the adapter.
    """
    from app.service.migration.intermediate import IntermediateParty
    from app.service.migration.vyapar_adapter import _COLUMN_MAP

    party_fields = set(IntermediateParty.model_fields.keys())
    internal_sentinels = {"_opening_balance", "_balance_type", "_party_type"}
    valid_targets = party_fields | internal_sentinels

    for header, target in _COLUMN_MAP.items():
        assert target in valid_targets, (
            f"_COLUMN_MAP[{header!r}] -> {target!r} doesn't match any "
            f"IntermediateParty field or known sentinel. "
            f"Valid targets: {sorted(valid_targets)}"
        )


def test_d3_column_map_keys_lowercase_unique() -> None:
    """Headers in _COLUMN_MAP must be lower-cased (the lookup in
    _read_header lower-cases the header text) and unique (a shadow
    mapping would silently drop coverage of one of the variants).
    """
    from app.service.migration.vyapar_adapter import _COLUMN_MAP

    for header in _COLUMN_MAP:
        assert header == header.lower(), (
            f"_COLUMN_MAP key {header!r} is not lower-cased; the lookup "
            "in _read_header lower-cases the header text before matching."
        )
    # dict can't have duplicate keys, but make the invariant explicit
    # so a future refactor (e.g. to a list[tuple]) keeps it.
    assert len(_COLUMN_MAP) == len(set(_COLUMN_MAP.keys()))


# ──────────────────────────────────────────────────────────────────────
# MIG-5: formula-injection sanitisation in extract_parties
# MIG-1: zip-bomb guard in _load_workbook
# MIG-2: malformed-file → AppValidationError in _load_workbook
# ──────────────────────────────────────────────────────────────────────


def _build_xlsx_with_formula_names() -> bytes:
    """Build an in-memory xlsx whose name/code cells start with formula triggers.

    The ``=EVIL()`` cell is written via the normal openpyxl value setter, which
    stores it as a formula.  When the adapter reads the file with
    ``data_only=True`` the cached formula result is absent (never evaluated),
    so openpyxl returns ``None`` for that cell — and the adapter skips the row
    entirely.  That leaves three parseable rows: ``+``, ``-``, and ``@``.

    Tests that call this helper must therefore expect **3** parties, not 4.
    """
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Parties")
    ws.append(["Party Name", "Code", "Type"])

    # Row 2: name starts with '=' — openpyxl stores this as a formula.
    # When read back with data_only=True the result is None (no cached value),
    # so the adapter drops the row.  This verifies the adapter doesn't crash on
    # formula cells and doesn't execute the formula.
    ws.cell(row=2, column=1).value = "=EVIL()"
    ws.cell(row=2, column=2).value = "SAFE-CODE"
    ws.cell(row=2, column=3).value = "Customer"

    # Row 3: name and code start with '+'.
    ws.cell(row=3, column=1).value = "+EVIL-NAME"
    ws.cell(row=3, column=2).value = "+EVIL-CODE"
    ws.cell(row=3, column=3).value = "Customer"

    # Row 4: trigger '-'.
    ws.cell(row=4, column=1).value = "-NEG"
    ws.cell(row=4, column=2).value = "-NEGCODE"
    ws.cell(row=4, column=3).value = "Customer"

    # Row 5: trigger '@'.
    ws.cell(row=5, column=1).value = "@SUM(A1)"
    ws.cell(row=5, column=2).value = "NORM"
    ws.cell(row=5, column=3).value = "Supplier"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extract_parties_stores_raw_formula_name_unchanged() -> None:
    """Party names with formula-trigger chars must be stored RAW — no ``'`` prefix.

    The ``=EVIL()`` row is dropped entirely (openpyxl returns ``None`` for an
    unevaluated formula cell with ``data_only=True``), leaving 3 parties.
    The remaining trigger-prefixed names (``+``, ``-``, ``@``) must be stored
    verbatim; the export layer is responsible for neutralising them at write time.
    """
    from app.service.migration import VyaparExcelAdapter

    adapter = VyaparExcelAdapter()
    parties = list(adapter.extract_parties(_build_xlsx_with_formula_names()))

    # =EVIL() row yields None name → skipped; 3 rows remain.
    assert len(parties) == 3, f"Expected 3 parties, got {len(parties)}: {[p.name for p in parties]}"
    names = {p.name for p in parties}
    # Raw values must be preserved — no apostrophe prefix injected at import.
    assert "+EVIL-NAME" in names, f"Raw '+EVIL-NAME' missing; got: {names}"
    assert "-NEG" in names, f"Raw '-NEG' missing; got: {names}"
    assert "@SUM(A1)" in names, f"Raw '@SUM(A1)' missing; got: {names}"
    for p in parties:
        assert not p.name.startswith("'"), f"Adapter must NOT sanitize on import; got {p.name!r}"


def test_extract_parties_stores_raw_formula_code_unchanged() -> None:
    """Party codes with formula-trigger chars must be stored RAW — no ``'`` prefix."""
    from app.service.migration import VyaparExcelAdapter

    adapter = VyaparExcelAdapter()
    parties = {p.name: p for p in adapter.extract_parties(_build_xlsx_with_formula_names())}

    # Row 3 has name "+EVIL-NAME" and code "+EVIL-CODE" — both stored raw.
    evil_plus = parties.get("+EVIL-NAME")
    assert evil_plus is not None, f"Party '+EVIL-NAME' not found; names: {list(parties)}"
    assert not evil_plus.code.startswith("'"), (
        f"Adapter must NOT sanitize code on import; got {evil_plus.code!r}"
    )
    assert evil_plus.code == "+EVIL-CODE", (
        f"Raw code must equal '+EVIL-CODE', got {evil_plus.code!r}"
    )


def test_validate_warns_on_formula_trigger_name() -> None:
    """``validate`` must emit a warning when a party name starts with a formula-trigger char.

    The raw value is stored unchanged; only the export layer neutralises it.
    A FORMULA_TRIGGER_IN_NAME reconciliation row lets the operator know the
    downstream export will prepend an apostrophe to that cell.
    """
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Parties")
    ws.append(["Party Name", "Type"])
    ws.append(["+Suspicious Name", "Customer"])
    ws.append(["Normal Name", "Customer"])
    out = BytesIO()
    wb.save(out)

    report = VyaparExcelAdapter().validate(out.getvalue())
    assert report.warnings >= 1, "Expected at least one warning for formula-trigger name"
    trigger_warnings = [r for r in report.rows if r.code == "FORMULA_TRIGGER_IN_NAME"]
    assert len(trigger_warnings) >= 1, (
        f"Expected FORMULA_TRIGGER_IN_NAME warning; got rows: {[r.code for r in report.rows]}"
    )


def test_zip_bomb_raises_app_validation_error(monkeypatch) -> None:  # type: ignore[no-untyped-def]
    """A zip whose uncompressed total exceeds the cap raises AppValidationError.

    Uses a REAL in-memory zip (no zipfile.ZipFile patch) and monkeypatches
    only the module-level cap constant to a small value so the test does not
    need to allocate gigabytes.  A spy on ``openpyxl.load_workbook`` asserts
    the guard fires BEFORE openpyxl is invoked — if the guard were absent the
    spy would record a call and the assertion would fail.
    """
    import io
    import zipfile
    from unittest.mock import patch

    import openpyxl as openpyxl_mod

    import app.service.migration.vyapar_adapter as adapter_mod
    from app.exceptions import AppValidationError
    from app.service.migration.vyapar_adapter import _load_workbook

    # Tiny cap: 40 KiB — well below our 50 KiB test zip.
    monkeypatch.setattr(adapter_mod, "_MAX_DECOMPRESSED_BYTES", 40 * 1024)

    # Real zip: 5 members x 10 KiB = 50 KiB uncompressed > 40 KiB cap.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for i in range(5):
            zf.writestr(f"chunk_{i}.txt", "A" * 10 * 1024)
    big_zip_bytes = buf.getvalue()

    # Spy on openpyxl.load_workbook — wraps the real function so the test
    # fails loudly if the guard lets through and openpyxl is reached.
    with patch.object(openpyxl_mod, "load_workbook", wraps=openpyxl_mod.load_workbook) as spy:
        with pytest.raises(AppValidationError, match="decompressed"):
            _load_workbook(big_zip_bytes)
        spy.assert_not_called()  # bomb guard must fire BEFORE openpyxl


def test_bad_zip_file_raises_app_validation_error() -> None:
    """Bytes that are not a zip (BadZipFile) become AppValidationError, not 500."""
    from app.exceptions import AppValidationError
    from app.service.migration.vyapar_adapter import _load_workbook

    with pytest.raises(AppValidationError, match=r"[Rr]eadable|[Ee]xcel|[Ww]orkbook"):
        _load_workbook(b"this is definitely not a zip file")


def test_invalid_xlsx_raises_app_validation_error() -> None:
    """A zip file that isn't a valid OOXML workbook raises AppValidationError."""
    import io
    import zipfile

    from app.exceptions import AppValidationError
    from app.service.migration.vyapar_adapter import _load_workbook

    # A valid zip that contains garbage — not an xlsx.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("garbage.txt", "not an ooxml workbook")
    fake_zip_bytes = buf.getvalue()

    with pytest.raises(AppValidationError, match=r"[Rr]eadable|[Ee]xcel|[Ww]orkbook"):
        _load_workbook(fake_zip_bytes)


def test_d3_party_type_map_values_are_party_kinds() -> None:
    """Every _PARTY_TYPE_MAP value must be a non-empty tuple of valid
    PartyKind literals.  Same regression guard as _COLUMN_MAP — keeps
    the spike runner's vocabulary printout truthful."""
    from typing import get_args

    from app.service.migration.intermediate import PartyKind
    from app.service.migration.vyapar_adapter import _PARTY_TYPE_MAP

    # PartyKind is a Literal — extract its allowed values from typing.
    valid_kinds = set(get_args(PartyKind))
    assert valid_kinds, "PartyKind must define at least one literal value"

    for src_value, kinds in _PARTY_TYPE_MAP.items():
        assert isinstance(kinds, tuple) and kinds, (
            f"_PARTY_TYPE_MAP[{src_value!r}] must be a non-empty tuple, got {kinds!r}"
        )
        for k in kinds:
            assert k in valid_kinds, (
                f"_PARTY_TYPE_MAP[{src_value!r}] contains {k!r} which is not "
                f"a valid PartyKind. Valid kinds: {sorted(valid_kinds)}"
            )
